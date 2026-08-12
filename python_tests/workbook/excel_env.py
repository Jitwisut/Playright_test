from __future__ import annotations

import os
import re
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_FILENAME = "Trainee BU3_ Manage Expopass.xlsx"
WORKSHEET_NAME = "Web Registration Online"

EMAIL_PATTERN = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
PASSWORD_PATTERN = re.compile(
    r"(?:password|รหัสผ่าน)\s*[:=]?\s*([^\s]+)",
    re.IGNORECASE,
)
URL_PATTERN = re.compile(r"https?://[^\s]+", re.IGNORECASE)


def discover_workbook_path() -> Path | None:
    configured = os.getenv("WORKBOOK_XLSX_PATH")
    if configured:
        path = Path(configured).expanduser()
        if not path.is_file():
            raise FileNotFoundError(f"WORKBOOK_XLSX_PATH does not exist: {path}")
        return path

    candidates = (
        Path.home() / "Downloads" / WORKBOOK_FILENAME,
        Path.cwd() / WORKBOOK_FILENAME,
    )
    return next((path for path in candidates if path.is_file()), None)


def _conference_credentials(value: object) -> tuple[str | None, str | None]:
    text = str(value or "")
    email_match = EMAIL_PATTERN.search(text)
    password_match = PASSWORD_PATTERN.search(text)
    email = email_match.group(0).strip() if email_match else None
    password = password_match.group(1).strip() if password_match else None
    return email, password


def _first_url(value: object) -> str | None:
    match = URL_PATTERN.search(str(value or ""))
    return match.group(0).rstrip(".,);]") if match else None


def load_workbook_environment() -> tuple[str, ...]:
    """Load safe runtime defaults from the workbook without logging secret values."""
    if os.getenv("WORKBOOK_AUTOLOAD_EXCEL", "1") == "0":
        return ()

    workbook_path = discover_workbook_path()
    if workbook_path is None:
        return ()

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Worksheet not found: {WORKSHEET_NAME}")

        sheet = workbook[WORKSHEET_NAME]
        conference_user, conference_password = _conference_credentials(sheet["F113"].value)
        invite_url = _first_url(sheet["F120"].value)
        defaults = {
            "WORKBOOK_CONFERENCE_USER": conference_user,
            "WORKBOOK_CONFERENCE_PASSWORD": conference_password,
            "WORKBOOK_INVITE_URL": invite_url,
        }
        loaded: list[str] = []
        for name, value in defaults.items():
            if value and name not in os.environ:
                os.environ[name] = value
                loaded.append(name)
        return tuple(loaded)
    finally:
        workbook.close()
